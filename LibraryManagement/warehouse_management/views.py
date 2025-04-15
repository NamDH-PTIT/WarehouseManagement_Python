import json
from datetime import datetime, timedelta
import random
from http.client import responses
from tarfile import data_filter

import openpyxl
from django.core.mail import send_mail
from django.db import transaction
from django.db.models import Q
from django.db.models.aggregates import Sum
from django.forms import model_to_dict
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from openpyxl.styles.builtins import total
from rest_framework import status
from unicodedata import category

from LibraryManagement import settings
from database.models import *
from warehouse_management.response.ApiResponse import ApiResponse


# Create your views here.
@require_http_methods(["GET"])
@require_http_methods(["GET"])
@csrf_exempt
def getAllProducts(request):
    products = list(Product.objects.raw("select * from database_product order by quantity asc"))
    page = int(request.GET.get('page', 1)) - 1
    low_stock_products = Product.objects.filter(quantity__lt=30)
    pagequantity = int(len(products) / 10)
    danhmuc = list(Product.objects.raw("SELECT category,id FROM `database_product` GROUP BY category"))
    if (len(products) % 10 != 0):
        pagequantity += 1
    return render(request, 'LibraryManagement/quanlyproduct.html',
                  {"products": products[page * 10:page * 10 + 10], "sanphamsaphet": low_stock_products,
                   "page": range(1, pagequantity + 1), "nowpage": page + 1, "danhmuc": danhmuc})


@require_http_methods(["GET"])
@csrf_exempt
def editProduct(request, product_id):
    # product = get_object_or_404(Product.objects.values(),code=product_id)
    product = Product.objects.get(code=product_id)  # Tìm sản phẩm theo code, nếu không có thì trả về 404
    category = Product.objects.raw("SELECT category,id FROM `database_product` GROUP BY category")
    request.session['product_id'] = product_id
    return render(request, "LibraryManagement/editProduct.html", {"product": product, "category": category})


@require_http_methods(["POST", "GET"])
@csrf_exempt
def addPhieuNhap(request):
    if request.method == "GET":
        ncc = NhaCungCap.objects.all()
        kho = Kho.objects.all()
        category = Product.objects.raw("SELECT category,id FROM `database_product` GROUP BY category")
        return render(request, "LibraryManagement/addPhieuNhap.html", {"ncc": ncc, "kho": kho, "category": category})

    ncc = request.POST.get("supplier")
    date = datetime.strptime(request.POST.get("date"), '%Y-%m-%d').date()
    warehouse = Kho.objects.get(code=request.POST.get("warehouse"))
    notes = request.POST.get("notes")
    productdata = json.loads(request.POST.get("productData"))
    totalprice = request.POST.get("totalAmount")
    try:
        with transaction.atomic():
            # nccRequest = data.get("nameNCC")
            nhacungcap = NhaCungCap.objects.get(code=ncc)
            # phieuNhapRequest = data.get("phieuNhap")
            phieuNhap = PhieuNhap.objects.create(
                # date=phieuNhapRequest.get("date"),
                # totalPrice=phieuNhapRequest.get("totalPrice"),
                # notes=phieuNhapRequest.get("notes"),
                # codeNCC=ncc
                date=date,
                totalPrice=int(totalprice),
                notes=notes,
                codeKho=warehouse,
                codeNCC=nhacungcap,

            )
            user = User.objects.filter(code=request.session.get('user_id')).first()
            if user is not None:
                log = Log.objects.create(
                    date=datetime.now(),
                    notes='tạo phiếu nhập ' + str(phieuNhap.code)
                )
                if user.role.name == 'Admin':
                    log.user = 'Admin'
                    phieuNhap.status = 'completed'

                else:
                    log.user = request.session.get('user_id')
                log.save()
                phieuNhap.save()
            for item in productdata:
                product = Product.objects.create(
                    nameProduct=item["name"],
                    category=item["type"],
                    quantity=item["quantity"],
                    sellingPrice=item["price"],
                    importPrice=item["price"],
                    codeKho=warehouse,
                    notes=item["notes"]
                )
                ChiTietPhieuNhap.objects.create(
                    codeProduct=product,
                    codePhieuNhap=phieuNhap,
                    quantity=item["quantity"],
                    importPrice=int(item["price"])
                )
            return redirect('/quanlynhaphang')
    except Exception as e:
        return JsonResponse({'e': str(e)})


@csrf_exempt
@require_http_methods(["POST"])
def addNCC(request):
    data = json.loads(request.body)
    nccRequest = data.get("nccInfo")
    if NhaCungCap.objects.filter(nameNCC=nccRequest.get("nameNCC")).exists():
        nccResponse = NhaCungCap.objects.get(nameNCC=nccRequest.get("nameNCC"))
        return JsonResponse({"message": "Nhà cung cấp đã tồn tại!", "model": model_to_dict(nccResponse)}, status=400)
    ncc = NhaCungCap.objects.create(
        nameNCC=nccRequest.get("nameNCC"),
        addressNCC=nccRequest.get("addressNCC"),
        code=nccRequest.get("code"),
        phoneNCC=nccRequest.get("phoneNCC"),
        emailNCC=nccRequest.get("emailNCC")
    )
    return JsonResponse({"message": "Thêm nhà cung cấp thành công", "model": model_to_dict(ncc)}, status=200)


@csrf_exempt
@require_http_methods(["POST"])
def addCustomer(request):
    name = request.POST.get("name")
    address = request.POST.get("address")
    phone = request.POST.get("phone")
    email = request.POST.get("email")
    status = request.POST.get("status")
    if Customer.objects.filter(email=email).exists():
        return render(request, 'LibraryManagement/notifi.html', {
            'message': 'đã tồn tại ',
            'back_url': '/quanlykhachhang/'  # Đổi endpoint này tùy vào bạn muốn quay về đâu
        })
        # return render(request, 'LibraryManagement/quanlykhachhang.html')
    customer = Customer.objects.create(
        name=name,
        phone=phone,
        email=email,
        address=address,
        status=status
    )
    return redirect('/quanlykhachhang/')


@csrf_exempt
@require_http_methods(["GET"])
def getCustomer(request, customer_code=None):
    if customer_code is None:
        customers = list(Customer.objects.values())
        return JsonResponse({"success": True, "customers": customers}, status=200)
    else:
        try:
            customer = Customer.objects.get(code=customer_code)
            return JsonResponse({"success": True, "customer": model_to_dict(customer)}, status=200)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=400)


@csrf_exempt
@require_http_methods(["GET"])
def getNCC(request, customer_id=None):
    if customer_id is None:
        nhaCungCap = list(NhaCungCap.objects.values())
        return JsonResponse({"success": True, "NhaCungCap": nhaCungCap}, status=200)
    else:
        try:
            customer = NhaCungCap.objects.get(id=customer_id)
            return JsonResponse({"success": True, "customer": model_to_dict(customer)}, status=200)
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)}, status=400)


# @csrf_exempt
# @require_http_methods(["POST"])
# @transaction.atomic
# def addPhieuXuat(request):
#     data = json.loads(request.body)
#     customerRequest = data.get("codeCustomer")
#
#     try:
#
#         customer = Customer.objects.get(code=customerRequest)
#
#         phieuXuatRequest = data.get("phieuXuat")
#         listChiTietPhieuXuat = data.get("chiTietPhieuXuat")
#
#         # **Bước 1: Kiểm tra toàn bộ danh sách trước khi tạo phiếu xuất**
#         for item in listChiTietPhieuXuat:
#             if not Product.objects.filter(code=item.get("codeProduct")).exists():
#                 return JsonResponse({"message": f"Không có sản phẩm {item.get('codeProduct')}", "success": False},
#                                     status=400)
#
#             product = Product.objects.get(code=item.get("codeProduct"))
#             if product.quantity < item.get("quantity"):
#                 return JsonResponse({
#                     "message": "Số lượng sản phẩm không đủ",
#                     "success": False,
#                     "model": model_to_dict(product)
#                 }, status=400)
#
#         # **Bước 2: Tạo phiếu xuất vì không có lỗi nào**
#         with transaction.atomic():
#             phieuXuat = PhieuXuat.objects.create(
#                 date=phieuXuatRequest.get("date"),
#                 totalPrice=phieuXuatRequest.get("totalPrice"),
#                 notes=phieuXuatRequest.get("notes"),
#                 customer=customer
#             )
#
#             # **Bước 3: Lưu chi tiết phiếu xuất**
#             for item in listChiTietPhieuXuat:
#                 product = Product.objects.get(code=item.get("codeProduct"))
#
#                 ChiTietPhieuXuat.objects.create(
#                     quantity=item.get("quantity"),
#                     sellingPrice=item.get("sellingPrice"),
#                     product=product,
#                     phieuXuat=phieuXuat
#                 )
#                 # **Cập nhật số lượng sản phẩm**
#                 product.quantity -= item.get("quantity")
#                 product.save()
#                 Log.objects.create(
#                     date=datetime.now(),
#                     notes="xử lý phiếu xuất" + phieuXuat.code
#                 )
#
#         return JsonResponse({"message": "Thêm phiếu xuất thành công", "success": True})
#
#     except Customer.DoesNotExist:
#         return JsonResponse({"message": "Khách hàng không tồn tại", "success": False}, status=400)
#     except Exception as e:
#         return JsonResponse({"success": False, "message": str(e)}, status=400)


@csrf_exempt
@require_http_methods(["POST"])
def updateProduct(request):
    product = Product.objects.get(code=request.session.get("product_id"))
    product.category = request.POST.get("productType")
    product.sellingPrice = request.POST.get("productPrice")
    product.image = request.FILES.get("productImage")
    product.nameProduct = request.POST.get("productName")
    product.save()
    return redirect("/getProduct/")


@csrf_exempt
@require_http_methods(["GET"])
def home_manager(request):
    low_stock_products = Product.objects.filter(quantity__lt=30)
    low_stock_products = low_stock_products.order_by("quantity")
    phieuxuat = PhieuXuat.objects.filter(status="completed")
    loinhuan = 0
    for item in phieuxuat:
        chitietpx = ChiTietPhieuXuat.objects.filter(phieuXuat=item)
        for x in chitietpx:
            tong = (x.sellingPrice - x.product.importPrice) * x.quantity
            loinhuan = tong
    doanhthu = PhieuXuat.objects.filter(status="completed").aggregate(Sum('totalPrice'))['totalPrice__sum']
    products = Product.objects.all()
    pxt = []
    for i in range(1, 13):
        pxt.append(PhieuXuat.objects.filter(date__month=i).count())
    phieuxuat = PhieuXuat.objects.filter(status="pending")

    log = Log.objects.order_by('-date')[:5]

    return render(request, 'LibraryManagement/home_manager.html',
                  {"pxt": pxt, "phieuxuat": phieuxuat, "soluongpx": phieuxuat.count(), "products": products.count(),
                   "log": log, "doanhthu": doanhthu, "loinhuan": loinhuan, "sanphamsaphet": low_stock_products})


@csrf_exempt
@require_http_methods(["POST"])
def export_products_excel(request):
    data = json.loads(request.body)
    category = data.get("category")
    status = data.get("status")
    sort = data.get("sort")
    search = data.get("search")
    products = Product.objects.all()

    if category:
        products = products.filter(category=category)
    if status:
        if status == "2":
            products = products.filter(quantity__gte=1)
        if status == "3":
            products = products.filter(quantity__lte=20)
        if status == "4":
            products = products.filter(quantity__lte=0)
    if search:
        products = products.filter(nameProduct__icontains=search)

    # Sắp xếp theo sort
    if sort == 'name-asc':
        products = products.order_by('nameProduct')
    if sort == 'name-desc':
        products = products.order_by('-nameProduct')
    if sort == "price-asc":
        products = products.order_by("sellingPrice")  # Sắp xếp tăng dần
    if sort == "price-desc":
        products = products.order_by("-sellingPrice")  # Sắp xếp giảm dần
    # Tạo file Excel mới
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Tiêu đề cột
    headers = ["ID", "Code", "Name", "Category", "Import Price", "Selling Price", "Quantity", "Notes"]
    ws.append(headers)

    # Lấy dữ liệu từ database và thêm vào Excel
    for product in products:
        ws.append([
            product.id, product.code, product.nameProduct, product.category,
            product.importPrice, product.sellingPrice, product.quantity, product.notes
        ])

    # Trả về file Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
    wb.save(response)

    return response


@csrf_exempt
@require_http_methods(["GET", "POST"])
def login(request):
    if request.method == "POST":
        email = request.POST.get("email")
        password = request.POST.get("password")
        user = User.objects.filter(email=email).first()
        if not user:
            return render(request, "LibraryManagement/login.html",
                          {'message': 'tài khoản hoặc mật khẩu không chính xác'})
        if user.password == password:
            request.session["user_id"] = user.code
            if user.role.name == "Admin":
                return redirect("/home_manager")
            elif user.role.name == "User":
                return redirect("/home")
            else:
                return redirect('/home_vc/')
        else:
            return render(request, "LibraryManagement/login.html",
                          {"message": "tài khoản hoặc mật khẩu không chính xác"})

    elif request.method == "GET":
        return render(request, "LibraryManagement/login.html")


def send_email(sub, mess, email):
    subject = sub
    message = mess
    from_email = settings.EMAIL_HOST_USER
    recipient_list = [email]
    send_mail(subject, message, from_email, recipient_list)


@csrf_exempt
@require_http_methods(["POST"])
def forgot(request):
    data = request.POST.get("email")
    if (User.objects.filter(email=data).first() is None):
        return JsonResponse({"success": False, "message": "email không tồn tại"}, status=200)
    otp = random.randint(100000, 999999)
    user = User.objects.filter(email=data).first()
    expiry_time = datetime.now() + timedelta(minutes=5)
    request.session["user_code"] = user.code
    request.session["otp"] = otp
    request.session["otp_expiry"] = expiry_time.strftime("%Y-%m-%d %H:%M:%S")
    send_email("mã OTP", "OTP : " + str(otp), data)
    return JsonResponse({"success": True}, status=200)


@csrf_exempt
@require_http_methods(["POST"])
def verify_otp(request):
    user_otp = request.POST.get("otp")  # OTP nhập từ người dùng
    session_otp = request.session.get("otp")  # OTP trong session
    otp_expiry = request.session.get("otp_expiry")  # Thời gian hết hạn
    if not session_otp or not otp_expiry:
        return JsonResponse({"message": "OTP không tồn tại hoặc đã hết hạn!", "success": False}, status=400)
    otp_expiry_time = datetime.strptime(otp_expiry, "%Y-%m-%d %H:%M:%S")
    if datetime.now() > otp_expiry_time:
        return JsonResponse({"message": "OTP đã hết hạn!", "success": False}, status=400)

    if str(user_otp) != str(session_otp):
        return JsonResponse({"message": "OTP không hợp lệ!", "success": False}, status=400)

    # Xóa OTP khỏi session sau khi xác minh thành công
    return JsonResponse({"message": "Xác minh thành công!", "success": True})


@csrf_exempt
@require_http_methods(["POST"])
def updatepassword(request):
    password = request.POST.get("password")
    user_code = request.session.get("user_code")
    User.objects.filter(code=user_code).update(password=password)


@csrf_exempt
@require_http_methods(["DELETE"])
def delete_product(request, product_code=None):
    product = get_object_or_404(Product, code=product_code)
    product.delete()
    return JsonResponse({"success": True})


@csrf_exempt
@require_http_methods(["GET"])
def filter_products(request):
    category = request.GET.get('category')
    sort = request.GET.get('sort')
    status = request.GET.get('status')
    search = request.GET.get('search')
    low_stock_products = Product.objects.filter(quantity__lt=30)
    low_stock_products = low_stock_products.order_by("quantity")
    # Lọc sản phẩm
    products = Product.objects.all()

    # Lọc theo category
    if category:
        products = products.filter(category=category)

    # Lọc theo status
    if status:
        if status == "2":
            products = products.filter(quantity__gte=1)
        if status == "3":
            products = products.filter(quantity__lte=20)
        if status == "4":
            products = products.filter(quantity__lte=0)

    # Lọc theo tìm kiếm
    if search:
        products = products.filter(nameProduct__icontains=search)

    # Sắp xếp theo sort
    if sort == 'name-asc':
        products = products.order_by('nameProduct')
    elif sort == 'name-desc':
        products = products.order_by('-nameProduct')
    elif sort == "price-asc":
        products = products.order_by("sellingPrice")  # Sắp xếp tăng dần
    elif sort == "price-desc":
        products = products.order_by("-sellingPrice")  # Sắp xếp giảm dần

    # Sắp xếp theo số lượng
    products = products.order_by('quantity')  # Sắp xếp giảm dần theo quantity

    # Trả về JSON hoặc render kết quả

    danhmuc = list(Product.objects.raw("SELECT category,id FROM `database_product` GROUP BY category"))
    return render(request, "LibraryManagement/quanlyproduct.html",
                  {"products": products, "danhmuc": danhmuc, 'sanphamsaphet': low_stock_products})


def getuser(request):
    user = list(User.objects.all())
    low_stock_products = Product.objects.filter(quantity__lt=30)
    return render(request, 'LibraryManagement/manage_account.html', {"user": user, "sanphamsaphet": low_stock_products})


@csrf_exempt
@require_http_methods(["GET"])
def user_filter(request):
    statusFilter = request.GET.get('status')
    sortFilter = request.GET.get('sort')
    search = request.GET.get('search')
    user_filter = (User.objects.all())
    if statusFilter == "active":
        user_filter = user_filter.filter(status='active')
    if statusFilter == "inactive":
        user_filter = user_filter.filter(status='inactive')
    if statusFilter == "pending":
        user_filter = user_filter.filter(status='')
    if sortFilter == "name-asc":
        user_filter = user_filter.order_by('name')
    if sortFilter == "name-desc":
        user_filter = user_filter.order_by('-name')
    if search is not None:
        user_filter = user_filter.filter(name__icontains=search)
    user = list(user_filter)
    return render(request, 'LibraryManagement/manage_account.html', {'user': user})


@csrf_exempt
@require_http_methods(["POST"])
def adduser(request):
    name = request.POST.get("name")
    email = request.POST.get("email")
    password = request.POST.get("password")
    status = request.POST.get("status")
    role = Role.objects.get(name=request.POST.get("role"))
    user_exists = User.objects.filter(
        Q(email=email)
    ).exists()
    if user_exists:
        message = "Tài khoản đã tồn tại!"
        return render(request, 'LibraryManagement/notifi.html', {
            'message': message,
            'back_url': '/getuser/'  # Đổi endpoint này tùy vào bạn muốn quay về đâu
        })
    str1 = str(uuid.uuid4())
    User.objects.create(
        code=str1,
        name=name,
        email=email,
        password=password,
        status=status,
        role=role
    )
    Log.objects.create(
        notes="tạo user " + str1,
        date=datetime.now()
    )

    return redirect('/getuser/')


@csrf_exempt
@require_http_methods(["POST"])
def export_user(request):
    data = json.loads(request.body)

    statusFilter = data.get("status")
    sortFilter = data.get("sort")
    search = data.get("search")
    user_filter = (User.objects.all())
    if statusFilter == "active":
        user_filter = user_filter.filter(status='active')
    if statusFilter == "inactive":
        user_filter = user_filter.filter(status='inactive')
    if statusFilter == "pending":
        user_filter = user_filter.filter(status='')
    if sortFilter == "name-asc":
        user_filter = user_filter.order_by('name')
    if sortFilter == "name-desc":
        user_filter = user_filter.order_by('-name')
    if search is not None:
        user_filter = user_filter.filter(name__icontains=search)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # Tiêu đề cột
    headers = ["ID", "Code", "Name", "address", "phone", "email", "password", "status", "role"]
    ws.append(headers)

    # Lấy dữ liệu từ database và thêm vào Excel
    for u in user_filter:
        ws.append([
            u.id, u.code, u.name, u.address, u.phone, u.email, u.password, u.status, u.role.name
        ])

    # Trả về file Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
    wb.save(response)
    return response


def getuserbyid(request, userid):
    user = User.objects.get(id=userid)
    return render(request, 'LibraryManagement/edituser.html', {'user': user})


@csrf_exempt
@require_http_methods(["POST"])
def updateuser(request):
    name_request = request.POST.get("name")
    email_request = request.POST.get("email")
    password_request = request.POST.get("password")
    phone_request = request.POST.get("phone")
    address_request = request.POST.get("address")
    role_request = request.POST.get("role")
    role = Role.objects.get(name=role_request)
    status = request.POST.get("status")

    user = User.objects.get(email=email_request)
    user.address = address_request
    user.name = name_request
    user.email = email_request
    user.password = password_request
    user.status = status
    user.role = role
    user.phone = phone_request
    user.save()
    return redirect('/getuser/')


def deleteuser(request, userid):
    user = User.objects.get(id=userid)
    user.delete()
    Log.objects.create(
        date=datetime.now(),
        notes="delete user " + userid,
    )
    return redirect('/getuser/')


@csrf_exempt
@require_http_methods(["POST", "GET"])
def addproduct(request):
    if request.method == "GET":
        category = Product.objects.raw("SELECT category,id FROM `database_product` GROUP BY category")
        kho = list(Kho.objects.all().values())
        return render(request, 'LibraryManagement/addproduct.html', {'category': category, 'kho': kho})
    elif request.method == "POST":
        nameProduct = request.POST.get("name")
        category = request.POST.get("type")
        importPrice = request.POST.get("importPrice")
        sellingPrice = request.POST.get("sellingPrice")
        quantity = request.POST.get("quantity")
        codeKho = Kho.objects.get(code=request.POST.get("codeKho"))
        notes = request.POST.get("notes")
        product = Product.objects.create(
            nameProduct=nameProduct,
            category=category,
            importPrice=importPrice,
            sellingPrice=sellingPrice,
            quantity=quantity,
            codeKho=codeKho,
            notes=notes,
        )
        Log.objects.create(
            date=datetime.now(),
            notes="thêm sản phẩm " + nameProduct + str(product.code),
        )
        return redirect('/getProduct/')


def warehouse_managment(request):
    codeKho = request.GET.get("kho")
    kho = list(Kho.objects.all().values())
    product = Product.objects.all()

    tenkho = ""
    if codeKho is None:
        tenkho = Kho.objects.first().description
        product = Product.objects.filter(codeKho=Kho.objects.first().code)
        product = product.order_by('quantity')  # Gán lại kết quả sau khi sắp xếp
    else:
        tenkho = Kho.objects.get(code=codeKho).description
        product = Product.objects.filter(codeKho=Kho.objects.get(code=codeKho))
        product = product.order_by('quantity')  # Gán lại kết quả sau khi sắp xếp
    return render(request, 'LibraryManagement/warehouse_managment.html',
                  {'kho': kho, 'product': product, 'tenkho': tenkho})


def exportkho(request):
    codeKho = request.GET.get('kho')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"
    product = Product.objects.all()

    if codeKho:
        product = Product.objects.filter(codeKho=Kho.objects.get(code=codeKho))
    else:
        product = Product.objects.filter(codeKho=Kho.objects.first().code)
    product = product.order_by('quantity')
    # Tiêu đề cột
    headers = ["id", "code", "nameProduct", "category", "importPrice", "sellingPrice", "quantity", "codeKho", "notes"]
    ws.append(headers)

    # Lấy dữ liệu từ database và thêm vào Excel
    for u in product:
        ws.append([
            u.id, u.code, u.nameProduct, u.category, u.importPrice, u.sellingPrice, u.quantity, u.codeKho.description,
            u.notes
        ])

    # Trả về file Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
    wb.save(response)
    return response


def chuyenkho(request, productCode=None):
    product = Product.objects.get(code=productCode)
    kho = Kho.objects.all()
    return render(request, 'LibraryManagement/chuyenkho.html', {'product': product, 'kho': kho})


@csrf_exempt
@require_http_methods(["POST"])
def chuyensp(request):
    data = json.loads(request.body)
    idProduct = data.get("id")
    product = Product.objects.get(code=idProduct)
    newWarehouse = str(data.get("warehouse"))
    kho = Kho.objects.get(description=newWarehouse)
    quantity = data.get("quantity")
    if (product.quantity - int(quantity) < 0):
        return (JsonResponse({'success': 'sai số lượng'}, status=500))
    if (kho == product.codeKho):
        Log.objects.create(
            date=datetime.now(),
            notes="sản phẩm vẫn giữ nguyên " + product.nameProduct + " " + str(product.code)
        )
        return JsonResponse({'success': True})
    else:
        product.quantity = product.quantity - int(quantity)
        product.save()
        Product.objects.create(
            nameProduct=product.nameProduct,
            category=product.category,
            importPrice=product.importPrice,
            sellingPrice=product.sellingPrice,
            quantity=quantity,
            codeKho=kho,
        )
        Log.objects.create(
            date=datetime.now(),
            notes="Sản phẩm được chuyển từ " + str(
                product.codeKho.description) + " tới " + newWarehouse + " số lượng " + quantity
        )
        return JsonResponse({'success': True})


def quanlynhaphang(request):
    ncc = NhaCungCap.objects.all()
    phieunhap = PhieuNhap.objects.all()
    kho = Kho.objects.all()
    trangthai = PhieuNhap.objects.raw('select id,status from database_phieunhap group by status')
    return render(request, 'LibraryManagement/quanlynhaphang.html',
                  {'ncc': ncc, 'phieunhap': phieunhap, 'kho': kho, 'trangthai': trangthai})


def quanlykhachhang(request):
    ncc = NhaCungCap.objects.all()
    user = Customer.objects.all().order_by('status')

    # Tạo dictionary từ customer_id -> số lượng phiếu xuất (k)
    soluong_raw = PhieuXuat.objects.raw(
        'SELECT id, customer_id, COUNT(*) as k FROM database_phieuxuat GROUP BY customer_id')
    soluong_dict = {row.customer_id: row.k for row in soluong_raw}

    # Gán số lượng vào từng user
    for u in user:
        u.soluong_phieuxuat = soluong_dict.get(u.code, 0)  # 'code' là khóa chính (primary key) của Customer

    # status giữ nguyên nếu bạn cần
    status = Customer.objects.raw('SELECT id, status FROM database_customer GROUP BY status')

    return render(request, 'LibraryManagement/quanlykhachhang.html', {
        'ncc': ncc,
        'user': user,
        'status': status,
    })


def quanlykhachhang_filter(request):
    vip = request.GET.get('status')
    sort = request.GET.get('sort')
    search = request.GET.get('search')
    customer = Customer.objects.all()
    customer = customer.order_by('status')
    if (vip == 'vip 1'):
        customer = customer.filter(status='vip 1')
    if (vip == 'vip 2'):
        customer = customer.filter(status='vip 2')
    if (vip == 'vip 3'):
        customer = customer.filter(status='vip 3')
    if (sort == 'name-asc'):
        customer = customer.order_by('name')
    if (sort == 'name-desc'):
        customer = customer.order_by('-name')
    if search is not None:
        customer = customer.filter(name__icontains=search)
    soluong_raw = PhieuXuat.objects.raw(
        'SELECT id, customer_id, COUNT(*) as k FROM database_phieuxuat GROUP BY customer_id')
    soluong_dict = {row.customer_id: row.k for row in soluong_raw}

    # Gán số lượng vào từng user
    for u in customer:
        u.soluong_phieuxuat = soluong_dict.get(u.code, 0)  # 'code' là khóa chính (primary key) của Customer

    return render(request, 'LibraryManagement/quanlykhachhang.html', {'user': customer})


@csrf_exempt
@require_http_methods(["POST"])
def customer_excel(request):
    data = json.loads(request.body)

    vip = data.get('vip')
    sort = data.get('sort')
    search = data.get('search')
    user_filter = (Customer.objects.all())
    if vip == "vip 1":
        user_filter = user_filter.filter(status='vip 1')
    if vip == "vip 2":
        user_filter = user_filter.filter(status='vip 2')
    if vip == "vip 3":
        user_filter = user_filter.filter(status='vip 3')
    if sort == "name-asc":
        user_filter = user_filter.order_by('name')
    if sort == "name-desc":
        user_filter = user_filter.order_by('-name')
    if search is not None:
        user_filter = user_filter.filter(name__icontains=search)

    soluong_raw = PhieuXuat.objects.raw(
        'SELECT id, customer_id, COUNT(*) as k FROM database_phieuxuat GROUP BY customer_id')
    soluong_dict = {row.customer_id: row.k for row in soluong_raw}

    # Gán số lượng vào từng user
    for u in user_filter:
        u.soluong_phieuxuat = soluong_dict.get(u.code, 0)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer"

    # Tiêu đề cột
    headers = ["ID", "Code", "Name", "address", "phone", "email", "status", "số đơn"]
    ws.append(headers)

    # Lấy dữ liệu từ database và thêm vào Excel
    for u in user_filter:
        ws.append([
            u.id, u.code, u.name, u.address, u.phone, u.email, u.status, u.soluong_phieuxuat
        ])

    # Trả về file Excel
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="products.xlsx"'
    wb.save(response)
    return response


def editcustomer(request, id=None):
    customer = Customer.objects.get(id=id)
    return render(request, 'LibraryManagement/editcustomer.html', {'user': customer})


@csrf_exempt
@require_http_methods(["POST"])
def updatecustomer(request):
    id = request.POST.get('id')
    name = request.POST.get('name')
    email = request.POST.get('email')
    phone = request.POST.get('phone')
    address = request.POST.get('address')
    status = request.POST.get('status')
    customer = Customer.objects.get(id=id)
    customer.name = name
    customer.email = email
    customer.phone = phone
    customer.address = address
    customer.status = status
    customer.save()
    return redirect('/quanlykhachhang/')


@csrf_exempt
@require_http_methods(["POST", "GET"])
def add_ncc(request):
    if request.method == "GET":
        return render(request, 'LibraryManagement/addncc.html')
    if request.method == "POST":
        nameNCC = request.POST.get('nameNCC')
        addressNCC = request.POST.get('addressNCC')
        phoneNCC = request.POST.get('phoneNCC')
        emailNCC = request.POST.get('emailNCC')
        noteNCC = request.POST.get('note')
        user_code = request.session.get('user_id', '')
        if NhaCungCap.objects.filter(emailNCC=emailNCC).exists():
            return render(request, 'LibraryManagement/notifi.html', {
                'message': 'Nhà cung cấp bị trùng thông tin email',
                'back_url': '/quanlynhaphang/'
            })
        NhaCungCap.objects.create(
            nameNCC=nameNCC,
            addressNCC=addressNCC,
            phoneNCC=phoneNCC,
            emailNCC=emailNCC,
            notes=noteNCC,
        )
        Log.objects.create(
            date=datetime.now(),
            notes='Tạo nhà cung cấp ' + str(nameNCC) + " " + str(emailNCC) + ' ' + str(user_code)
        )
        return render(request, 'LibraryManagement/quanlyxuathang.html')


def quanlyxuathang(request):
    trangthai = PhieuXuat.objects.raw('select id,status from database_phieuxuat group by status')
    kh = Customer.objects.all()
    kho = Kho.objects.all()
    phieuxuat = PhieuXuat.objects.all()
    return render(request, 'LibraryManagement/quanlyxuathang.html',
                  {'trangthai': trangthai, 'kh': kh, 'kho': kho, 'phieuxuat': phieuxuat})


@csrf_exempt
@require_http_methods(["POST", "GET"])
def thongbao(request):
    if request.method == "GET":
        users = User.objects.exclude(role__name='Admin')
        return render(request, 'LibraryManagement/thongbao.html', {'users': users})
    elif request.method == "POST":
        content = request.POST.get('content')
        recipientType = request.POST.get('recipientType')
        user = request.POST.get('user_code')
        if recipientType == 'specific':
            Log.objects.create(
                date=datetime.now(),
                notes=content,
                user=user
            )
        if recipientType == 'all':
            Log.objects.create(
                date=datetime.now(),
                notes=content,
                user=recipientType
            )
        return render(request, 'LibraryManagement/thongbao.html')


@csrf_exempt
@require_http_methods(["POST", "GET"])
def chitietphieunhap(request, id):
    phieunhap = PhieuNhap.objects.get(code=id)
    chitietphieunhap = ChiTietPhieuNhap.objects.filter(codePhieuNhap=phieunhap)
    ncc=Kho.objects.all()
    return render(request, 'LibraryManagement/chitietphieunhap.html',
                  {'phieunhap': phieunhap, 'chitietphieunhap': chitietphieunhap,'ncc': ncc})


@csrf_exempt
@require_http_methods(["POST", "GET"])
def update_phieunhap(request):
    data = json.loads(request.body)
    status = data.get('status')
    code = data.get('code')
    date = data.get('date')
    ncc = NhaCungCap.objects.get(nameNCC=data.get('supplier'))
    warehouse = Kho.objects.get(description=data.get('warehouse'))
    notes = data.get('note')
    phieunhap = PhieuNhap.objects.get(code=code)
    phieunhap.status = status
    phieunhap.date = date_obj = datetime.strptime(date, "%B %d, %Y, midnight")
    phieunhap.notes = notes
    phieunhap.codeNCC = ncc
    phieunhap.codeKho = warehouse
    product = data.get('products')
    chitietphieunhap = ChiTietPhieuNhap.objects.filter(codePhieuNhap=phieunhap)
    total=0
    for item in product:
        for x in chitietphieunhap:
            if item['code'] == x.codeProduct.code:

                x.codeProduct.quantity=item['quantity']
                x.quantity = item['quantity']
                total=total+(item['quantity']*item['importPrice'])
                x.importPrice=item['importPrice']
                x.codeProduct.importPrice=item['importPrice']
                x.codeProduct.save()
                x.save()

    phieunhap.totalPrice=total
    phieunhap.save()
    Log.objects.create(
        date=datetime.now(),
        notes='Admin cập nhật phiếu nhập '+str(code),
    )
    return JsonResponse({'status': True})
@csrf_exempt
@require_http_methods(["POST", "GET"])
def baocao(request):
    if request.method == "GET":
        return render(request,'LibraryManagement/nvbaocao.html')
    content=request.POST.get('content')
    send_email('báo cáo',content,'hainam26112003@gmail.com')
    Log.objects.create(
        date=datetime.now(),
        notes='nhân viên gửi mail báo cáo'
    )
    return render(request,'LibraryManagement/nvbaocao.html')
@csrf_exempt
@require_http_methods(["POST", "GET"])
def logout(request):
    request.session.flush()
    return redirect('/')
def user_login(request):
    user_id=request.session["user_id"]
    user=User.objects.filter(code=user_id)
    products = Product.objects.all()
    productsaphet = Product.objects.filter(quantity__lt=20)
    phieunhap = PhieuNhap.objects.order_by("-date")[:12]
    nhiemvu = Log.objects.filter(Q(user='all') | Q(user='user')).order_by("-date")[:4]
    phieuxuat = PhieuXuat.objects.order_by("-date")[:12]
    phieuxuatpending = PhieuXuat.objects.filter(status="pending").count()
    kho = Product.objects.raw(
        'SELECT id,codeKho_id, COUNT(*) AS so_luong_san_pham  FROM database_product  GROUP BY codeKho_id')
    total = 0
    hoatdong1 = Log.objects.all().order_by("-date")[:1]
    hoatdong2 = Log.objects.all().order_by("-date")[1:2]
    hoatdong3 = Log.objects.all().order_by("-date")[2:3]
    hoatdong4 = Log.objects.all().order_by("-date")[3:4]
    for item in products:
        total = total + (item.quantity * item.importPrice)
    return render(request, "LibraryManagement/user.html",
                  {'user': user, 'total': total, 'nhiemvu': nhiemvu, 'phieunhap': phieunhap,
                   'phieuxuat': phieuxuat, 'products': products.count(),
                   'productsaphet': productsaphet.count(), 'phieuxuatpending': phieuxuatpending, 'kho': kho,
                   'hoatdong1': hoatdong1, 'hoatdong2': hoatdong2, 'hoatdong3': hoatdong3,
                   'hoatdong4': hoatdong4})


@csrf_exempt
@require_http_methods(["POST", "GET"])
def home_vc(request):
    if(request.method == "GET"):
        return render(request,'LibraryManagement/nvvanchuyen.html')
@csrf_exempt
@require_http_methods(["POST", "GET"])
def update_vc(request,id):
    print('àafsaf')
    return render(request,'LibraryManagement/nvvanchuyen.html')
@transaction.atomic
@csrf_exempt
@require_http_methods(["POST", "GET"])
def addphieuxuat(request):
    if(request.method == "GET"):
        customer = Customer.objects.all()
        products = Product.objects.all()
        products = products.filter(quantity__gt=0)
        return render(request, 'LibraryManagement/addphieuxuat.html', {'customer': customer, 'products': products})
    date = datetime.strptime(request.POST.get("date"), '%Y-%m-%d').date()
    customer = Customer.objects.filter(code=request.POST.get("customer")).first()
    notes = request.POST.get("notes")
    status = request.POST.get("status")
    total = request.POST.get("totalPrice")
    productData = json.loads(request.POST.get("productData"))
    phieu_xuat = PhieuXuat.objects.create(
        date=date,
        totalPrice=total,
        notes=notes,
        status=status,
        customer=customer,
    )
    for item in productData:
        sanpham = Product.objects.filter(code=item['code']).first()
        sanpham.quantity = sanpham.quantity - int(item['quantity'])
        sanpham.save()
        chitietphieuxuat = ChiTietPhieuXuat.objects.create(
            quantity=int(item['quantity']),
            sellingPrice=sanpham.sellingPrice,
            phieuXuat=phieu_xuat,
            product=sanpham,
        )
    soluongdon = PhieuXuat.objects.filter(customer=customer).count()
    if soluongdon < 5:
        customer.status = 'vip 3'
    elif soluongdon < 10:
        customer.status = 'vip 2'
    elif soluongdon < 20:
        customer.status = 'vip 1'
    else:
        customer.status = 'vip 0'
    customer.save()

    return redirect('/quanlyxuathang/')



def chitietphieuxuat(request,id):
    phieuxuat=PhieuXuat.objects.filter(code=id).first()
    customer=Customer.objects.all()
    products = Product.objects.all()
    chitietphieuxuat=ChiTietPhieuXuat.objects.filter(phieuXuat=phieuxuat)
    return render(request,'LibraryManagement/chitietphieuxuat.html',{'phieuxuat': phieuxuat,'customer':customer,'products':products,'chitietphieuxuat':chitietphieuxuat})